# import thư viện
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import os
import smtplib
import shutil
import ssl
from email.message import EmailMessage
from openpyxl import load_workbook
# Thư mục tải hóa đơn
name_folder_download = "HoaDonDienTu"
# Hàm đọc mã tra cứu và mật khẩu từ file Excel
def doc_du_lieu_tu_excel(file_excel):
    ma_tra_cuu_list = []
    password = None
    wb = load_workbook(filename=file_excel)
    sheet = wb.active
    for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
        ma_tra_cuu, pass_connect = row
        if ma_tra_cuu:
            ma_tra_cuu_list.append(ma_tra_cuu.strip())
        if pass_connect and not password:
            password = pass_connect.strip()
    return ma_tra_cuu_list, password
# Hàm tự động mở web và tra cứu hóa đơn
def input_hoa_don_auto(ma_tra_cuu):
    os.makedirs(name_folder_download, exist_ok=True)
    options = webdriver.ChromeOptions()
    options.add_argument('--start-maximized')
    prefs = {
        "download.default_directory": os.path.abspath(name_folder_download),
        "download.prompt_for_download": False,
        "plugins.always_open_pdf_externally": True
    }
    options.add_experimental_option("prefs", prefs)
    driver = webdriver.Chrome(options=options)
    driver.get("https://www.meinvoice.vn/tra-cuu")
    try:
        WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.NAME, "txtCode"))
        )
        input_code_element = driver.find_element(By.NAME, "txtCode")
        input_code_element.send_keys(ma_tra_cuu)
        print("Đang tra cứu mã:", ma_tra_cuu)

        btn_search_element = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.ID, "btnSearchInvoice"))
        )
        btn_search_element.click()
        print("Đã nhấn tra cứu")

        WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.XPATH, "//span[contains(text(),'Tải hóa đơn')]"))
        )
        print("Hóa đơn đã sẵn sàng để tải")
        return driver
    except Exception as error:
        print("Mã tra cứu sai:", error)
        driver.quit()
        return None
# Hàm tải hóa đơn PDF
def tai_hoa_don(driver):
    try:
        download_btn = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CLASS_NAME, "download"))
        )
        driver.execute_script("arguments[0].scrollIntoView(true);", download_btn)
        time.sleep(1)
        driver.execute_script("arguments[0].click();", download_btn)
        print("Đã nhấn nút chính 'Tải hóa đơn'")

        pdf_option = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CLASS_NAME, "txt-download-pdf"))
        )
        driver.execute_script("arguments[0].click();", pdf_option)
        print("Đã chọn tải hóa đơn dạng PDF")
        time.sleep(10)
        return True
    except Exception as e:
        print("Lỗi khi tải hóa đơn PDF:", e)
        return False
# Hàm tìm file PDF mới nhất
def search_file_new(folder):
    print("Đang chờ file tải xong...")
    timeout = 20
    start_time = time.time()
    while time.time() - start_time < timeout:
        files = os.listdir(folder)
        full_paths = [os.path.join(folder, f) for f in files if f.endswith(".pdf") and not f.endswith(".crdownload")]
        if full_paths:
            file_path = max(full_paths, key=os.path.getctime)
            print("Đã tìm thấy file:", file_path)
            return file_path
        time.sleep(1)
    print("Không tìm thấy file nào.")
    return None
# Hàm gửi file hóa đơn qua email
def gui_file_ve_gmail(file_path, gmail_gui, mat_khau_app, gmail_nhan):
    subject = "Hóa đơn điện tử từ meInvoice"
    body = "Đây là file hóa đơn bạn vừa tra cứu tự động."
    msg = EmailMessage()
    msg["From"] = gmail_gui
    msg["To"] = gmail_nhan
    msg["Subject"] = subject
    msg.set_content(body)
    with open(file_path, "rb") as f:
        file_data = f.read()
        file_name = os.path.basename(file_path)
    msg.add_attachment(file_data, maintype="application", subtype="octet-stream", filename=file_name)

    context = ssl.create_default_context()
    with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
        server.login(gmail_gui, mat_khau_app)
        server.send_message(msg)
        print(f"Đã gửi file '{file_name}' về Gmail:", gmail_nhan)
# Chương trình chính
if __name__ == "__main__":
    print("Bắt đầu tra cứu và gửi hóa đơn...\n")
    # Đọc từ Excel
    file_excel = "Ma_tra_cuu.xlsx"
    ma_tra_cuu_list, password_connect_email = doc_du_lieu_tu_excel(file_excel)
    email_gui = "hp859869@gmail.com"
    email_nhan = "hp859869@gmail.com"
    for ma_tra_cuu in ma_tra_cuu_list:
        print(f"Đang xử lý mã: {ma_tra_cuu}")
        driver = input_hoa_don_auto(ma_tra_cuu)
        if driver:
            if tai_hoa_don(driver):
                driver.quit()
                file_moi = search_file_new(name_folder_download)
                if file_moi:
                    new_file_path = os.path.join(name_folder_download, f"{ma_tra_cuu}.pdf")
                    shutil.move(file_moi, new_file_path)
                    print("Đã đổi tên file thành:", new_file_path)
                    gui_file_ve_gmail(new_file_path, email_gui, password_connect_email, email_nhan)
                else:
                    print("Không tìm thấy file nào để gửi.")
            else:
                print("Tải hóa đơn thất bại.")
        else:
            print("Không thể mở trình duyệt hoặc nhập mã.")
        print("--------------------------------------------------\n")
